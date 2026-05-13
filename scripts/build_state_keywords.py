"""Generate state_keywords_analysis.xlsx — top headline events and recurring
themes per US state, plus a per-state outlier-summary sheet.

Mirrors the logic in index.html's state popup so the Excel and the website
always show the same data:

  1. Tag merging (data/tag_config.json — applied upstream by build_data.py /
     patch_beats.py to article subject arrays). Includes auto-titlecase of
     ALL-CAPS tags with abbrev preservation.
  2. Generic-tag filters from tag_config.json.
  3. Overrepresentation score = (tag freq in state / state articles) /
     (tag freq in corpus / corpus articles). Skips tags absent from corpus
     (avoids the divide-by-1 inflation that caused the AGRICULTURE bug).
  4. Headline-event classification by tag NAME structure ONLY:
       (a) Curated `headline_event_tags` exact match
       (b) `headline_event_patterns` substring match
       (c) Tag contains a single 4-digit year not part of a range
     We do NOT use statistical year-spike detection — it conflates
     "topic that became prominent in a single year" (e.g. In Vitro
     Fertilization in 2024) with "specific dated event" (e.g. Hurricane Ian
     2022). Headline events should be incidents, not subjects.
  5. Top 10 candidates → split: all top-10 headline events go to the
     headline column; up to 7 of the remainder (by score) go to recurring.

Run from the project root after build_data.py / patch_beats.py:
    python3 scripts/build_state_keywords.py

Produces: state_keywords_analysis.xlsx (in repo root) with two sheets:
  - State Outsize Subjects: full list per state, with score / % / type
  - Per-State Summary: min/max overrepresentation and % per state, for
    spotting weak vs. strong themes (use to set inclusion thresholds for
    a 50-state map deliverable).
"""

import json
import os
import re
from collections import Counter, defaultdict

PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(PROJECT_DIR, 'data')

with open(os.path.join(DATA_DIR, 'tag_config.json')) as f:
    TAG_CONFIG = json.load(f)

GENERIC_ALWAYS = set(TAG_CONFIG.get('generic_subjects_always_filter', []))
GENERIC_PREFIXES = tuple(TAG_CONFIG.get('generic_prefixes_always_filter', []))
STATE_GENERIC = set(TAG_CONFIG.get('state_coverage_generic_subjects', []))
STATE_GENERIC_PREFIXES = tuple(TAG_CONFIG.get('state_coverage_generic_prefixes', []))
HEADLINE_TAGS = set(TAG_CONFIG.get('headline_event_tags', []))
HEADLINE_PATTERNS = TAG_CONFIG.get('headline_event_patterns', [])

# Single year in tag name, NOT part of a range like "1939-45" or "2003- ".
# Range marker: '-' or '–' immediately after the year (with optional whitespace).
_YEAR_RE = re.compile(r'\b(19|20)\d{2}\b(?!\s*[-–])')

# Year-range tag where the START year is 1990 or later. Catches multi-year
# wars/conflicts that span our coverage period — e.g. 'Iraq War (2003-11)',
# 'Afghanistan War (2001- )', 'Israel-Gaza War (2023- )'. Distinguishes
# real ranges from identifier-style suffixes ('Coronavirus (2019-nCoV)')
# by requiring the dash to be followed by a digit or close-paren, not a
# letter. Historical eras like 'World War II (1939-45)' or 'Civil Rights
# Movement (1954-68)' don't match because they start before 1990 — they
# stay as recurring (retrospective coverage spread across our years).
_YEAR_RANGE_RE = re.compile(r'\b(199\d|20\d\d)\s*[-–]\s*[\d)]')

# Correction articles inflate state subject counts with topics unrelated to
# the state's actual beat (a correction's subject reflects what was being
# corrected, not what the state itself covers). Detect via subject tag,
# section name, or URL pattern — any one is sufficient.
_CORR_URL_RE = re.compile(r'/(c-)?corrections?-|/pageoneplus/corrections-')

def is_correction_article(a):
    sb = a.get('sb') or []
    if 'Correction Stories' in sb:
        return True
    if (a.get('s') or '') == 'Corrections':
        return True
    return bool(_CORR_URL_RE.search(a.get('u') or ''))

# Standing-feature listing headlines (event calendars, briefs, real-estate
# listings, art-review roundups) inflate state subject scores by manufacturing
# phony recurring themes — e.g. CT's 304 'Culture (Arts)' articles were 97%
# 'Events in Connecticut' calendar items.
_LISTING_EXACT = set(TAG_CONFIG.get('state_listing_headlines_exact', []))
_LISTING_PREFIXES = tuple(TAG_CONFIG.get('state_listing_headline_prefixes', []))
_LISTING_KICKERS = set(TAG_CONFIG.get('state_listing_kickers_exact', []))

def is_listing_article(a):
    h = a.get('h') or ''
    if h in _LISTING_EXACT:
        return True
    if any(h.startswith(p) for p in _LISTING_PREFIXES):
        return True
    return (a.get('k') or '') in _LISTING_KICKERS


SINGLE_VENUE = set(TAG_CONFIG.get('single_venue_tags', []))

def is_state_junk_tag(tag):
    if tag in STATE_GENERIC or tag in GENERIC_ALWAYS:
        return True
    if tag in SINGLE_VENUE:
        return True
    return (any(tag.startswith(p) for p in STATE_GENERIC_PREFIXES) or
            any(tag.startswith(p) for p in GENERIC_PREFIXES))


def is_headline_event(tag):
    """A tag is a headline event when its NAME structurally indicates a
    specific dated incident — not when its coverage clusters in time."""
    if tag in HEADLINE_TAGS:
        return True
    for p in HEADLINE_PATTERNS:
        if p in tag:
            return True
    if _YEAR_RE.search(tag):
        return True
    if _YEAR_RANGE_RE.search(tag):
        return True
    return False


YEAR_BURST_THRESHOLD = 0.85   # 2-consecutive-year share

def is_year_burst_state(years_counter):
    """For a (state, tag) year distribution, return True if >=85% of the
    coverage is concentrated in any pair of consecutive years. Catches
    state-specific event surges that the tag NAME doesn't reveal — e.g.
    'Serial Murders' in MD/VA (DC Sniper 2002-3), 'Fourteenth Amendment'
    in CO (Trump ballot case 2023-4)."""
    if not years_counter:
        return False
    total = sum(years_counter.values())
    if total < 5:
        return False
    sorted_yrs = sorted(years_counter.keys())
    max_pair = 0
    for i in range(len(sorted_yrs) - 1):
        if int(sorted_yrs[i + 1]) == int(sorted_yrs[i]) + 1:
            pair = years_counter[sorted_yrs[i]] + years_counter[sorted_yrs[i + 1]]
            if pair > max_pair:
                max_pair = pair
    return max_pair / total > YEAR_BURST_THRESHOLD


def load_articles():
    articles = []
    for fn in sorted(os.listdir(DATA_DIR)):
        if fn.startswith('articles_') and fn.endswith('.json'):
            with open(os.path.join(DATA_DIR, fn)) as fh:
                articles.extend(json.load(fh))
    return articles


def analyze(articles):
    """Return {state: {'headline': [...], 'recurring': [...]}}."""
    # Score denominator: U.S./New York-section coverage only. The whole-corpus
    # baseline used previously systematically inflated state disproportion
    # scores because it included articles (Sports, Arts, Style, etc.) that
    # couldn't plausibly carry geographic tags. Using just the state-eligible
    # pool gives an apples-to-apples comparison: how often does a tag appear
    # in U.S./New York coverage of *state X* versus that coverage overall.
    corpus_freq = Counter()
    total_corpus = 0
    for art in articles:
        if (art.get('s') or '') not in ('U.S.', 'New York'):
            continue
        total_corpus += 1
        seen = set()
        for tag in (art.get('sb') or []):
            if tag in GENERIC_ALWAYS or any(tag.startswith(p) for p in GENERIC_PREFIXES):
                continue
            if tag in seen:
                continue
            corpus_freq[tag] += 1
            seen.add(tag)

    state_articles = defaultdict(list)
    for art in articles:
        for st in (art.get('st') or []):
            state_articles[st].append(art)

    out = {}
    for state in sorted(state_articles.keys()):
        arts = state_articles[state]
        arts = [a for a in arts if not is_correction_article(a)
                and not is_listing_article(a)]
        state_total = len(arts)
        if state_total < 50:
            continue

        tag_counts = Counter()
        # Per-tag year distribution within this state — used to detect
        # year-burst topics that are state-specific event surges.
        tag_years = defaultdict(Counter)
        for a in arts:
            seen = set()
            yr = (a.get('d') or '')[:4]
            for tag in (a.get('sb') or []):
                if is_state_junk_tag(tag) or tag in seen:
                    continue
                tag_counts[tag] += 1
                if yr:
                    tag_years[tag][yr] += 1
                seen.add(tag)

        min_count = max(3, int(state_total * 0.005))
        scored = []
        for tag, cnt in tag_counts.items():
            if cnt < min_count:
                continue
            cf = corpus_freq.get(tag)
            if not cf:
                continue
            score = (cnt / state_total) / (cf / total_corpus) if total_corpus else cnt
            scored.append({
                'tag': tag,
                'count': cnt,
                'pct': round(cnt / state_total * 100, 2),
                'score': round(score, 1),
            })
        scored.sort(key=lambda x: -x['score'])

        # Helper: is this (state, tag) event-bound by either name structure
        # or a state-specific year burst?
        # Headline classification: tag-name structure only. Year-burst
        # statistical clustering catches multi-year recurring topics
        # (election cycles, ongoing conflicts) and miscalls them as events.
        # Same approach as the live state popup, which never used year-burst.
        def _is_event(tag):
            return is_headline_event(tag)

        # Headline events: all top-10 entries that classify as event-driven.
        # Recurring subjects: remaining top-10, padded to 5 from beyond
        # top-10 if needed (>= 10x national average required for padding).
        RECURRING_MIN, RECURRING_MAX, RECURRING_PAD_FLOOR = 5, 7, 10.0
        top10 = scored[:10]
        headline = [t for t in top10 if _is_event(t['tag'])]
        recurring = [t for t in top10 if not _is_event(t['tag'])]
        if len(recurring) < RECURRING_MIN:
            for t in scored[10:]:
                if len(recurring) >= RECURRING_MIN:
                    break
                if _is_event(t['tag']):
                    continue
                if t['score'] < RECURRING_PAD_FLOOR:
                    break
                recurring.append(t)
        recurring = recurring[:RECURRING_MAX]
        out[state] = {'headline': headline, 'recurring': recurring}
    return out


def write_excel(results):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.formatting.rule import ColorScaleRule
    except ImportError:
        print('ERROR: openpyxl not installed. pip3 install openpyxl')
        return False

    wb = Workbook()

    # ── Sheet 1: full per-state list ────────────────────────────────────
    ws = wb.active
    ws.title = 'State Outsize Subjects'
    ws.append(['State', 'Type', 'Subject', 'Articles', '% of state', 'Score'])
    for cell in ws[1]:
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        cell.font = Font(bold=True)

    for state in sorted(results.keys()):
        items = results[state]
        first = True
        for kind in ('headline', 'recurring'):
            for item in items[kind]:
                ws.append([
                    state if first else '',
                    'Headline event' if kind == 'headline' else 'Recurring',
                    item['tag'],
                    item['count'],
                    item['pct'] / 100.0,  # written as fraction so % format applies
                    item['score'],
                ])
                first = False

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = '0.00%'

    last_row = ws.max_row
    if last_row > 1:
        ws.conditional_formatting.add(
            f'E2:E{last_row}',
            ColorScaleRule(start_type='min', start_color='FFFFFF',
                           end_type='max', end_color='66BB66'),
        )
        ws.conditional_formatting.add(
            f'F2:F{last_row}',
            ColorScaleRule(start_type='min', start_color='FFFFFF',
                           end_type='max', end_color='E57373'),
        )

    # ── Sheet 2: per-state summary (outlier detection) ──────────────────
    # Key question for the 50-state map: is each shown theme a SERIOUS
    # outlier, or is it borderline? This sheet shows, per state and per
    # column (recurring vs. headline), the min/max score and min/max %
    # of state coverage among the items we'd display. If a state's
    # weakest recurring theme is, say, 2x with 0.4% of state articles,
    # that's the one to question.
    ws2 = wb.create_sheet('Per-State Summary')
    headers = [
        'State',
        '# recurring shown', 'Recurring max score', 'Recurring min score',
        'Recurring max %', 'Recurring min %',
        '# headline shown', 'Headline max score', 'Headline min score',
        'Headline max %', 'Headline min %',
    ]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='center')

    def _stats(items, key):
        if not items:
            return ('', '')
        vals = [it[key] for it in items]
        return (max(vals), min(vals))

    for state in sorted(results.keys()):
        rec = results[state]['recurring']
        head = results[state]['headline']
        rec_score_max, rec_score_min = _stats(rec, 'score')
        rec_pct_max, rec_pct_min = _stats(rec, 'pct')
        h_score_max, h_score_min = _stats(head, 'score')
        h_pct_max, h_pct_min = _stats(head, 'pct')
        ws2.append([
            state,
            len(rec),
            rec_score_max if rec else '',
            rec_score_min if rec else '',
            (rec_pct_max / 100.0) if rec else '',
            (rec_pct_min / 100.0) if rec else '',
            len(head),
            h_score_max if head else '',
            h_score_min if head else '',
            (h_pct_max / 100.0) if head else '',
            (h_pct_min / 100.0) if head else '',
        ])

    ws2.column_dimensions['A'].width = 22
    for col in 'BCDEFGHIJK':
        ws2.column_dimensions[col].width = 14
    # Format % columns
    for row in ws2.iter_rows(min_row=2):
        for col_idx in (5, 6, 10, 11):  # 1-indexed: E, F, J, K
            cell = row[col_idx - 1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.00%'

    last_row2 = ws2.max_row
    if last_row2 > 1:
        # Color scale: weak themes (low min score) stand out as light, so use
        # white→red where red = high score. We want WEAK to look pale.
        ws2.conditional_formatting.add(
            f'C2:D{last_row2}',
            ColorScaleRule(start_type='min', start_color='FFFFFF',
                           end_type='max', end_color='E57373'),
        )
        ws2.conditional_formatting.add(
            f'E2:F{last_row2}',
            ColorScaleRule(start_type='min', start_color='FFFFFF',
                           end_type='max', end_color='66BB66'),
        )

    out_path = os.path.join(PROJECT_DIR, 'outputs', '2026-05-top-keyword',
                            '2026-05-12-us-state-tweets', '-Tweets',
                            'state_keywords_analysis.xlsx')
    wb.save(out_path)
    print(f'  Saved {out_path}')
    return True


if __name__ == '__main__':
    print('Loading articles…')
    arts = load_articles()
    print(f'  {len(arts):,} articles')
    print('Analyzing state coverage…')
    results = analyze(arts)
    print(f'  {len(results)} states scored')
    print('Writing Excel…')
    write_excel(results)
    print('Done.')

"""
Fetch NYT bio 'thumbLarge' photos for the top-100 (co)bylined 2025 reporters.

Output:
  data/bio_photos/<slug>.<ext>     one image per reporter (when available)
  data/bio_photos/manifest.json    name / slug / rank / bio_url / image_url / filename / status

Source list: /Users/tedalcorn/Desktop/NYT_top100_2025.xlsx
"""

import json
import os
import re
import time
import urllib.request

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(ROOT, "data", "bio_photos")
MANIFEST = os.path.join(OUT_DIR, "manifest.json")

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122 Safari/537.36")


def load_targets():
    from openpyxl import load_workbook
    xlsx = "/Users/tedalcorn/Desktop/NYT_top100_2025.xlsx"
    wb = load_workbook(xlsx)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(values_only=True):
        rank, name, psec, w25, a25, ssec = r[:6]
        if isinstance(rank, int) and name:
            rows.append({"rank": rank, "name": name, "primary_section": psec,
                         "words_2025": w25, "articles_2025": a25})
    bios = json.load(open(os.path.join(ROOT, "data", "author_bios.json")))
    for r in rows:
        b = bios.get(r["name"])
        r["slug"] = b["slug"] if b else None
        r["bio_url"] = b["url"] if b else None
    return rows


def fetch(url):
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read()


def find_thumb_large(html, slug):
    # Prefer author-<slug>-thumbLarge[-v{n}].png for custom bio photos.
    pat = re.compile(
        rf'https://static\d+\.nyt\.com/images/[^"\s]+/author-{re.escape(slug)}-thumbLarge[^"\s]*?\.(?:png|jpg|jpeg)',
        re.IGNORECASE,
    )
    m = pat.search(html)
    if m:
        return m.group(0)
    # Fallback: any thumbLarge under an author-<slug> directory.
    pat2 = re.compile(
        rf'https://static\d+\.nyt\.com/images/[^"\s]+author-{re.escape(slug)}[^"\s]*?thumbLarge[^"\s]*?\.(?:png|jpg|jpeg)',
        re.IGNORECASE,
    )
    m = pat2.search(html)
    if m:
        return m.group(0)
    # Columnists: images/YYYY/MM/DD/opinion/<slug>/<slug>-thumbLarge.png (no "author-" prefix).
    pat3 = re.compile(
        rf'https://static\d+\.nyt\.com/images/[^"\s]+/{re.escape(slug)}-thumbLarge[^"\s]*?\.(?:png|jpg|jpeg)',
        re.IGNORECASE,
    )
    m = pat3.search(html)
    if m:
        return m.group(0)
    # Last-resort fallback: first thumbLarge on the page. Bio pages usually have
    # the headshot as the only (or first) thumbLarge asset, even when the slug
    # doesn't match (middle-initial stripped, underscores vs hyphens, name order).
    pat4 = re.compile(
        r'https://static\d+\.nyt\.com/images/[^"\s]+-thumbLarge[^"\s]*?\.(?:png|jpg|jpeg)',
        re.IGNORECASE,
    )
    m = pat4.search(html)
    return m.group(0) if m else None


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    targets = load_targets()
    print(f"Fetching bio photos for {len(targets)} reporters...")

    manifest = []
    for i, t in enumerate(targets, 1):
        name, slug, bio_url = t["name"], t["slug"], t["bio_url"]
        entry = {"rank": t["rank"], "name": name, "slug": slug,
                 "primary_section": t["primary_section"], "bio_url": bio_url,
                 "image_url": None, "filename": None, "status": None}

        if not slug or not bio_url:
            entry["status"] = "no_bio_url"
            manifest.append(entry)
            print(f"  [{i:3d}] {name} — no bio url")
            continue

        try:
            html = fetch(bio_url).decode("utf-8", errors="replace")
        except Exception as e:
            entry["status"] = f"fetch_error: {e}"
            manifest.append(entry)
            print(f"  [{i:3d}] {name} — fetch error: {e}")
            continue

        img_url = find_thumb_large(html, slug)
        if not img_url:
            entry["status"] = "no_thumb_large_found"
            manifest.append(entry)
            print(f"  [{i:3d}] {name} — no thumbLarge on bio page")
            continue

        ext = os.path.splitext(img_url.split("?")[0])[1].lower() or ".png"
        fname = f"{slug}{ext}"
        fpath = os.path.join(OUT_DIR, fname)
        try:
            data = fetch(img_url)
            with open(fpath, "wb") as f:
                f.write(data)
            entry["image_url"] = img_url
            entry["filename"] = fname
            entry["status"] = "ok"
            print(f"  [{i:3d}] {name} -> {fname}")
        except Exception as e:
            entry["status"] = f"image_fetch_error: {e}"
            print(f"  [{i:3d}] {name} — image fetch error: {e}")

        manifest.append(entry)
        time.sleep(0.4)

    with open(MANIFEST, "w") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)

    ok = sum(1 for m in manifest if m["status"] == "ok")
    print(f"\nDone. {ok}/{len(manifest)} images saved to {OUT_DIR}")
    print(f"Manifest: {MANIFEST}")


if __name__ == "__main__":
    main()

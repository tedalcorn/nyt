"""Generate country_keywords_analysis.xlsx — top headline events and
recurring subjects per country (top 50 by NYT World-section coverage).

Mirrors scripts/build_state_keywords.py but scoped to countries. Filters,
scoring, and headline/recurring split match the country popup in index.html
(see the "Outsize subjects" block in renderCountryDetail).

Run:
    python3 scripts/build_country_keywords.py

Produces: country_keywords_analysis.xlsx in the same -documents folder as
the state version, with three sheets:
  1. "Country Outsize Subjects" — every recurring + headline tag shown,
     one row per (country, tag) with count, % of country coverage, score
  2. "Top 5 Recurring per Country" — compact lookup view, one row per
     country with the five recurring subjects laid out left-to-right
  3. "Per-Country Summary" — min/max score and min/max % among shown items
"""
import json
import os
import re
import sys
from collections import Counter, defaultdict

PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(PROJECT_DIR, 'data')

with open(os.path.join(DATA_DIR, 'tag_config.json')) as fh:
    TAG_CONFIG = json.load(fh)

GENERIC_ALWAYS = set(TAG_CONFIG.get('generic_subjects_always_filter', []))
GENERIC_PREFIXES = tuple(TAG_CONFIG.get('generic_prefixes_always_filter', []))
WORLD_GENERIC = set(TAG_CONFIG.get('world_coverage_generic_subjects', []))
WORLD_GENERIC_SUBSTRS = TAG_CONFIG.get('world_coverage_generic_substrings', [])
HEADLINE_TAGS = set(TAG_CONFIG.get('headline_event_tags', []))
HEADLINE_PATTERNS = TAG_CONFIG.get('headline_event_patterns', [])

# Match country popup logic in index.html line 6184:
WORLD_PREFIXES = GENERIC_PREFIXES + ('Content Type:', 'Content type:', 'Vis-')

TOP_N_COUNTRIES = 50
RECURRING_MIN = 5
RECURRING_MAX = 7
RECURRING_PAD_FLOOR = 10.0
YEAR_BURST_THRESHOLD = 0.66  # same as state version


CORRECTION_URL_RE = re.compile(r'/(c-)?corrections?-|/pageoneplus/corrections-')

YEAR_BARE_RE = re.compile(r'\b(19|20)\d{2}\b(?!\s*[-–])')
YEAR_RANGE_RE = re.compile(r'\b(199\d|20\d\d)\s*[-–]\s*[\d)]')


def is_correction_article(a):
    if 'Correction Stories' in (a.get('sb') or []):
        return True
    if (a.get('s') or '') == 'Corrections':
        return True
    return bool(CORRECTION_URL_RE.search(a.get('u') or ''))


def is_world_junk_tag(tag, country):
    if tag in GENERIC_ALWAYS:
        return True
    if tag in WORLD_GENERIC:
        return True
    if any(tag.startswith(p) for p in WORLD_PREFIXES):
        return True
    if any(sub in tag for sub in WORLD_GENERIC_SUBSTRS):
        return True
    if tag.lower() == country.lower():
        return True
    return False


def is_headline_event(tag):
    if tag in HEADLINE_TAGS:
        return True
    for p in HEADLINE_PATTERNS:
        if p in tag:
            return True
    if YEAR_BARE_RE.search(tag):
        return True
    if YEAR_RANGE_RE.search(tag):
        return True
    return False


def is_year_burst(years_counter):
    """A tag is event-driven if 66%+ of its mentions cluster in 2 adjacent
    years — picks up state/country-specific event surges that don't carry
    a year in the tag name (Vieques 'Navies' 2001, Boston Marathon-related
    tags 2013, etc.)."""
    if not years_counter:
        return False
    total = sum(years_counter.values())
    if total < 5:
        return False
    sorted_yrs = sorted(years_counter.keys())
    max_pair = 0
    for i in range(len(sorted_yrs) - 1):
        if int(sorted_yrs[i + 1]) == int(sorted_yrs[i]) + 1:
            pair = years_counter[sorted_yrs[i]] + years_counter[sorted_yrs[i + 1]]
            if pair > max_pair:
                max_pair = pair
    return max_pair / total > YEAR_BURST_THRESHOLD


def load_world_articles():
    arts = []
    for fn in sorted(os.listdir(DATA_DIR)):
        if not (fn.startswith('articles_') and fn.endswith('.json')):
            continue
        with open(os.path.join(DATA_DIR, fn)) as fh:
            arts.extend(json.load(fh))
    return arts


def analyze(arts):
    # Score denominator: World-section coverage only. The whole-corpus baseline
    # used elsewhere on the dashboard systematically overstates disproportion
    # because it includes articles (Sports, Arts, Style, etc.) that couldn't
    # plausibly carry foreign-affairs tags. Using just World gives an
    # apples-to-apples comparison: how often does a tag appear in NYT World
    # coverage of *country X* versus NYT World coverage overall.
    world = [a for a in arts if (a.get('s') or '') == 'World']
    corpus_freq = Counter()
    total_corpus = 0
    for a in world:
        total_corpus += 1
        seen = set()
        for tag in (a.get('sb') or []):
            if tag in GENERIC_ALWAYS:
                continue
            if any(tag.startswith(p) for p in GENERIC_PREFIXES):
                continue
            if tag in seen:
                continue
            corpus_freq[tag] += 1
            seen.add(tag)

    # Regions to skip — these appear as glocations but are continents/areas,
    # not individual countries we'd map. Keeps them out of country-cards and
    # the eventual regional map analyses without removing them from the
    # World tab (where they remain valid glocations).
    SKIP_REGIONS = {'Africa', 'Europe'}

    # Country coverage by gn = normalized glocations.
    country_total = Counter()
    for a in world:
        for loc in (a.get('gn') or []):
            if loc in SKIP_REGIONS:
                continue
            country_total[loc] += 1
    top_countries = [c for c, _ in country_total.most_common(TOP_N_COUNTRIES)]

    out = {}
    for country in top_countries:
        country_arts = [a for a in world
                        if country in (a.get('gn') or [])
                        and not is_correction_article(a)]
        n_country = len(country_arts)
        if n_country < 30:  # small-sample countries get no result
            continue

        tag_counts = Counter()
        tag_years = defaultdict(Counter)
        for a in country_arts:
            yr = (a.get('d') or '')[:4]
            seen = set()
            for tag in (a.get('sb') or []):
                if is_world_junk_tag(tag, country):
                    continue
                if tag in seen:
                    continue
                tag_counts[tag] += 1
                if yr:
                    tag_years[tag][yr] += 1
                seen.add(tag)

        min_count = max(3, int(n_country * 0.01))
        scored = []
        for tag, cnt in tag_counts.items():
            if cnt < min_count:
                continue
            cf = corpus_freq.get(tag)
            if not cf:
                continue
            score = (cnt / n_country) / (cf / total_corpus) if total_corpus else cnt
            scored.append({
                'tag': tag,
                'count': cnt,
                'pct': round(cnt / n_country * 100, 2),
                'score': round(score, 1),
            })
        scored.sort(key=lambda x: -x['score'])

        # Headline classification: tag-name structure only. We deliberately
        # do NOT use year-burst statistical clustering — a multi-year recurring
        # topic (Bollywood, Wagner mercenaries, Han Chinese ethnic dynamics,
        # Indian election cycles, broadcast-decency, the Iraq-era POW/looting
        # coverage) often clusters in adjacent years and gets miscalled as an
        # event. Tag-name structure (year in name, year-range pattern, explicit
        # list) is the cleaner signal — same approach as the live country popup.
        def _event(tag, t_years):
            return is_headline_event(tag)

        top10 = scored[:10]
        headline = [t for t in top10 if _event(t['tag'], tag_years.get(t['tag']))]
        recurring = [t for t in top10 if not _event(t['tag'], tag_years.get(t['tag']))]

        if len(recurring) < RECURRING_MIN:
            for t in scored[10:]:
                if len(recurring) >= RECURRING_MIN:
                    break
                if _event(t['tag'], tag_years.get(t['tag'])):
                    continue
                if t['score'] < RECURRING_PAD_FLOOR:
                    break
                recurring.append(t)
        recurring = recurring[:RECURRING_MAX]

        out[country] = {
            'n_country': n_country,
            'headline': headline,
            'recurring': recurring,
        }

    return out


def write_excel(results):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.formatting.rule import ColorScaleRule
    except ImportError:
        print('ERROR: openpyxl not installed. pip3 install openpyxl', file=sys.stderr)
        return False

    wb = Workbook()

    # Stable display order: descending by total World-section articles
    ordered = sorted(results.items(), key=lambda kv: -kv[1]['n_country'])

    # ── Sheet 1: full per-country list ─────────────────────────────────
    ws = wb.active
    ws.title = 'Country Outsize Subjects'
    ws.append(['Country', 'Country articles', 'Type', 'Subject',
               'Articles', '% of country', 'Score (disproportion)'])
    for cell in ws[1]:
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        cell.font = Font(bold=True)

    for country, items in ordered:
        first = True
        for kind in ('headline', 'recurring'):
            for item in items[kind]:
                ws.append([
                    country if first else '',
                    items['n_country'] if first else '',
                    'Headline event' if kind == 'headline' else 'Recurring',
                    item['tag'],
                    item['count'],
                    item['pct'] / 100.0,
                    item['score'],
                ])
                first = False

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 18
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = '0.00%'

    last_row = ws.max_row
    if last_row > 1:
        ws.conditional_formatting.add(
            f'F2:F{last_row}',
            ColorScaleRule(start_type='min', start_color='FFFFFF',
                           end_type='max', end_color='66BB66'),
        )
        ws.conditional_formatting.add(
            f'G2:G{last_row}',
            ColorScaleRule(start_type='min', start_color='FFFFFF',
                           end_type='max', end_color='E57373'),
        )

    # ── Sheet 2: compact top-5-recurring lookup ────────────────────────
    ws2 = wb.create_sheet('Top 5 Recurring per Country')
    cols2 = ['Country', 'Country articles']
    for i in range(1, 6):
        cols2 += [f'#{i} Subject', f'#{i} Articles', f'#{i} %', f'#{i} Score']
    ws2.append(cols2)
    for cell in ws2[1]:
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='center')

    for country, items in ordered:
        row = [country, items['n_country']]
        for i in range(5):
            if i < len(items['recurring']):
                t = items['recurring'][i]
                row += [t['tag'], t['count'], t['pct'] / 100.0, t['score']]
            else:
                row += ['', '', '', '']
        ws2.append(row)

    ws2.column_dimensions['A'].width = 26
    ws2.column_dimensions['B'].width = 14
    # Subject columns wider, stats narrower
    for i in range(5):
        base = 3 + i * 4  # 1-indexed col index for #i Subject
        ws2.column_dimensions[chr(ord('A') + base - 1)].width = 38
        ws2.column_dimensions[chr(ord('A') + base)].width = 10
        ws2.column_dimensions[chr(ord('A') + base + 1)].width = 10
        ws2.column_dimensions[chr(ord('A') + base + 2)].width = 12
    # Format % cells
    for row in ws2.iter_rows(min_row=2):
        for i in range(5):
            cell = row[2 + i * 4 + 2]  # 0-indexed: column with "%"
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.00%'

    # ── Sheet 3: per-country summary (outlier detection) ──────────────
    ws3 = wb.create_sheet('Per-Country Summary')
    headers = [
        'Country', 'Country articles',
        '# recurring shown', 'Recurring max score', 'Recurring min score',
        'Recurring max %', 'Recurring min %',
        '# headline shown', 'Headline max score', 'Headline min score',
        'Headline max %', 'Headline min %',
    ]
    ws3.append(headers)
    for cell in ws3[1]:
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='center')

    def _stats(items, key):
        if not items:
            return ('', '')
        vals = [it[key] for it in items]
        return (max(vals), min(vals))

    for country, items in ordered:
        rec, head = items['recurring'], items['headline']
        rec_score_max, rec_score_min = _stats(rec, 'score')
        rec_pct_max, rec_pct_min = _stats(rec, 'pct')
        h_score_max, h_score_min = _stats(head, 'score')
        h_pct_max, h_pct_min = _stats(head, 'pct')
        ws3.append([
            country, items['n_country'],
            len(rec),
            rec_score_max if rec else '',
            rec_score_min if rec else '',
            (rec_pct_max / 100.0) if rec else '',
            (rec_pct_min / 100.0) if rec else '',
            len(head),
            h_score_max if head else '',
            h_score_min if head else '',
            (h_pct_max / 100.0) if head else '',
            (h_pct_min / 100.0) if head else '',
        ])

    ws3.column_dimensions['A'].width = 26
    ws3.column_dimensions['B'].width = 14
    for col in 'CDEFGHIJKL':
        ws3.column_dimensions[col].width = 14
    for row in ws3.iter_rows(min_row=2):
        for col_idx in (6, 7, 11, 12):
            cell = row[col_idx - 1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.00%'

    out_path = os.path.join(
        PROJECT_DIR, '-documents', 'top-keyword', '-Tweets',
        'country_keywords_analysis.xlsx',
    )
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f'  Saved {out_path}')
    return True


if __name__ == '__main__':
    print('Loading articles…')
    arts = load_world_articles()
    print(f'  {len(arts):,} articles')
    print('Analyzing country coverage…')
    results = analyze(arts)
    print(f'  {len(results)} countries scored')
    print('Writing Excel…')
    write_excel(results)
    print('Done.')

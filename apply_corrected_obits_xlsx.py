#!/usr/bin/env python3
"""Apply corrections from /Users/tedalcorn/Desktop/Corrected obits.xlsx.

Two outputs:
1. Patches data/obituaries.json in-place (surgical, no rebuild needed).
2. Appends entries to OBIT_OVERRIDES / OBIT_SPLITS / NON_OBIT_URLS in
   build_obituaries.py so a future rebuild preserves the corrections.
"""
import json
import re
import shutil
import urllib.parse
from openpyxl import load_workbook
from pathlib import Path

ROOT = Path(__file__).parent
XLSX = Path('/Users/tedalcorn/Desktop/Corrected obits.xlsx')
OBIT_JSON = ROOT / 'data' / 'obituaries.json'
BUILD_PY = ROOT / 'build_obituaries.py'


def to_path(u):
    return urllib.parse.urlparse(u or '').path


def read_xlsx():
    wb = load_workbook(XLSX, data_only=False)
    ws = wb.active
    out = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        note = (row[0].value or '').strip() if row[0].value else ''
        name = (row[2].value or '').strip() if row[2].value else ''
        age = row[3].value
        gender = row[4].value
        link = ''
        for cell in row:
            if cell.hyperlink:
                link = cell.hyperlink.target
                break
        if not link:
            continue
        out.append({'note': note, 'name': name, 'age': age,
                    'gender': gender, 'link': link})
    return out


def categorize(rows):
    overrides = {}
    splits = {}
    non_obit = set()
    for r in rows:
        p = to_path(r['link'])
        note = r['note']
        name = r['name']
        gender = r['gender']
        age = r['age']
        if 'Not an obit' in note:
            non_obit.add(p)
            continue
        if note.startswith('Link to two bios:') or note.startswith('Link to two obits:'):
            if 'bogdanoff' in p:
                splits[p] = [
                    {'name': 'Grichka Bogdanoff', 'age': 72, 'gender': 'M', 'gender_src': 'manual'},
                    {'name': 'Igor Bogdanoff', 'age': 72, 'gender': 'M', 'gender_src': 'manual'},
                ]
            elif 'ilse-nathan' in p or 'nathan' in p:
                splits[p] = [
                    {'name': 'Ilse Nathan', 'age': 98, 'gender': 'F', 'gender_src': 'manual'},
                    {'name': 'Ruth Siegler', 'age': 95, 'gender': 'F', 'gender_src': 'manual'},
                ]
            continue
        od = {}
        if gender in ('M', 'F', 'X'):
            od['gender'] = gender
            od['gender_src'] = 'manual'
        if name.startswith('Correction:'):
            od['name'] = name[len('Correction:'):].strip()
        if isinstance(age, int) and age > 0:
            od['age'] = age
        if od:
            overrides[p] = od
    return overrides, splits, non_obit


def patch_obit_json(overrides, splits, non_obit):
    obs = json.load(open(OBIT_JSON))
    by_path = {}
    for o in obs:
        by_path.setdefault(to_path(o.get('url')), []).append(o)

    # 1. Apply overrides
    over_applied = 0
    for p, od in overrides.items():
        if p not in by_path:
            print(f'  WARN: override path not found: {p}')
            continue
        for o in by_path[p]:
            for k, v in od.items():
                # For age: only overwrite if missing OR if explicitly given by user
                if k == 'age' and o.get('age') and o['age'] == v:
                    continue
                o[k] = v
            # When name overridden, sync display_name
            if 'name' in od:
                o['display_name'] = od['name']
            over_applied += 1
    print(f'  Overrides applied: {over_applied}')

    # 2. Apply splits — replace one row with N rows
    new_obs = []
    split_applied = 0
    for o in obs:
        p = to_path(o.get('url'))
        if p in splits:
            base = dict(o)
            base.pop('name', None)
            base.pop('display_name', None)
            base.pop('age', None)
            base.pop('gender', None)
            base.pop('gender_src', None)
            for sub in splits[p]:
                row = dict(base)
                row.update(sub)
                row['display_name'] = sub.get('name', row.get('display_name'))
                new_obs.append(row)
            split_applied += 1
            continue
        new_obs.append(o)
    print(f'  Splits applied: {split_applied}')

    # 3. Remove non-obits
    removed = 0
    final = []
    for o in new_obs:
        if to_path(o.get('url')) in non_obit:
            removed += 1
            continue
        final.append(o)
    print(f'  Non-obits removed: {removed}')

    # Backup + write
    shutil.copy(OBIT_JSON, str(OBIT_JSON) + '.bak')
    with open(OBIT_JSON, 'w') as f:
        json.dump(final, f, ensure_ascii=False, indent=2)
    print(f'  Wrote {len(final)} obits to {OBIT_JSON}')


def patch_build_py(overrides, splits, non_obit):
    src = BUILD_PY.read_text()

    # 1. Append to OBIT_OVERRIDES (insert before its closing `}` at top level)
    # Find the OBIT_OVERRIDES dict
    m = re.search(r'OBIT_OVERRIDES = \{', src)
    if not m:
        raise SystemExit('OBIT_OVERRIDES not found')
    # walk forward to matching }
    depth = 0
    i = m.end() - 1  # at the {
    while i < len(src):
        c = src[i]
        if c == '{':
            depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0:
                close_idx = i
                break
        i += 1
    # Insert before close_idx
    additions = ['\n    # ---- 2026-04-27 manual review (Corrected obits.xlsx, 130 rows) ----']
    for p, od in sorted(overrides.items()):
        items = ', '.join(f'{k!r}: {v!r}' for k, v in od.items())
        additions.append(f'    {p!r}: {{{items}}},')
    src = src[:close_idx] + '\n'.join(additions) + '\n' + src[close_idx:]

    # 2. Append to OBIT_SPLITS
    m = re.search(r'OBIT_SPLITS = \{', src)
    if not m:
        raise SystemExit('OBIT_SPLITS not found')
    depth = 0
    i = m.end() - 1
    while i < len(src):
        c = src[i]
        if c == '{':
            depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0:
                close_idx = i
                break
        i += 1
    additions = ['\n    # ---- 2026-04-27 manual review (Corrected obits.xlsx) ----']
    for p, lst in sorted(splits.items()):
        additions.append(f'    {p!r}: [')
        for sub in lst:
            items = ', '.join(f'{k!r}: {v!r}' for k, v in sub.items())
            additions.append(f'        {{{items}}},')
        additions.append('    ],')
    src = src[:close_idx] + '\n'.join(additions) + '\n' + src[close_idx:]

    # 3. Append to NON_OBIT_URLS
    m = re.search(r'NON_OBIT_URLS = \{', src)
    if not m:
        raise SystemExit('NON_OBIT_URLS not found')
    depth = 0
    i = m.end() - 1
    while i < len(src):
        c = src[i]
        if c == '{':
            depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0:
                close_idx = i
                break
        i += 1
    additions = ['\n    # ---- 2026-04-27 manual review (Corrected obits.xlsx) ----']
    for p in sorted(non_obit):
        additions.append(f'    {p!r},')
    src = src[:close_idx] + '\n'.join(additions) + '\n' + src[close_idx:]

    shutil.copy(BUILD_PY, str(BUILD_PY) + '.bak')
    BUILD_PY.write_text(src)
    print(f'  Patched {BUILD_PY}')


def main():
    rows = read_xlsx()
    print(f'Read {len(rows)} rows from xlsx')
    overrides, splits, non_obit = categorize(rows)
    print(f'  overrides: {len(overrides)}, splits: {len(splits)}, non_obit: {len(non_obit)}')
    print('\nPatching data/obituaries.json...')
    patch_obit_json(overrides, splits, non_obit)
    print('\nPatching build_obituaries.py...')
    patch_build_py(overrides, splits, non_obit)
    print('\nDone.')


if __name__ == '__main__':
    main()
